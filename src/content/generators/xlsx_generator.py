"""
Excel spreadsheet generator
"""

import json
import logging
import os
from typing import Optional, Dict, List, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from src.content.generators.base_generator import BaseGenerator
from src.config.language_utils import get_translation

class XlsxGenerator(BaseGenerator):
    """Generator for Excel spreadsheets (.xlsx)"""
    
    # Excel has a 31 character limit for sheet names
    SHEET_NAME_MAX_LENGTH = 31
    
    # Characters not allowed in Excel sheet names
    INVALID_SHEET_CHARS = [':', '\\', '/', '?', '*', '[', ']']
    
    def _validate_sheet_name(self, name: str) -> str:
        """
        Validate and sanitize Excel sheet names.
        
        Args:
            name: Original sheet name
            
        Returns:
            Sanitized sheet name that meets Excel's requirements
        """
        # Trim whitespace
        sanitized = name.strip()
        
        # Remove invalid characters
        for char in self.INVALID_SHEET_CHARS:
            sanitized = sanitized.replace(char, '_')
            
        # Truncate to maximum length
        if len(sanitized) > self.SHEET_NAME_MAX_LENGTH:
            logging.warning(f"Sheet name '{name}' exceeds maximum length of {self.SHEET_NAME_MAX_LENGTH} characters, truncating.")
            sanitized = sanitized[:self.SHEET_NAME_MAX_LENGTH]
            
        # Excel doesn't allow empty sheet names
        if not sanitized:
            sanitized = "Sheet"
            
        return sanitized
    
    def generate(self, directory: str, filename: str, description: str,
                industry: str, language: str, role: Optional[str] = None,
                date_range_str: Optional[str] = None) -> bool:
        """
        Generate Excel spreadsheet content.
        
        Args:
            directory: Target directory for the file
            filename: Name of the file to create
            description: Description of the file content
            industry: Industry context
            language: Language to use
            role: Specific role within the industry (optional)
            date_range_str: Date range information (optional)
            
        Returns:
            True if file was successfully created, False otherwise
        """
        if not XLSX_AVAILABLE:
            logging.error("openpyxl package is not available. Please install it with: pip install openpyxl")
            return False
            
        file_path = self.get_file_path(directory, filename)
        
        # Load system message from language resources
        try:
            # Get system message for structured output
            system_message = get_translation("content_generation.xlsx_generation", language)
                
        except Exception as e:
            logging.error(f"Failed to load system message from language resources for '{language}': {e}")
            return False
        
        # Create prompt for spreadsheet content
        prompt = self.create_prompt(description, industry, language, role, "Excel spreadsheet", date_range_str)
        if not prompt:
            logging.error(f"Failed to create prompt for {filename} with language '{language}'")
            return False
        
        # Generate structured content using LLM
        content = self.llm_client.get_json_completion(
            prompt=prompt,
            system_prompt=system_message,
            max_attempts=3,
            language=language
        )
        
        if not content or "sheets" not in content:
            logging.error(f"Failed to generate valid spreadsheet content for {filename}")
            return False
        
        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Process each sheet
            for sheet_data in content["sheets"]:
                # Get sheet name and validate it
                original_sheet_name = sheet_data.get("name", "Sheet")
                sheet_name = self._validate_sheet_name(original_sheet_name)
                
                # Create sheet
                sheet = wb.create_sheet(title=sheet_name)
                
                # Add headers
                headers = sheet_data.get("headers", [])
                for col_idx, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data rows
                rows = sheet_data.get("data", [])
                for row_idx, row_data in enumerate(rows, start=2):
                    for col_idx, value in enumerate(row_data, start=1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col_idx, _ in enumerate(headers, start=1):
                    # Approximate width calculation
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            
            # Save workbook
            wb.save(file_path)
            return True
        except Exception as e:
            logging.error(f"Failed to create Excel document {filename}: {e}")
            return False 

    def generate_excel_data(self, file_description: str, industry: str, folder_path: str, language: str, role: Optional[str] = None) -> Dict:
        """
        Generate excel file data with the provided parameters.
        
        Args:
            file_description: Description of the file purpose
            industry: Industry context
            folder_path: Path of the folder containing the file 
            language: Language to use
            role: User role context
            
        Returns:
            Dictionary containing cells data
        """
        role_context = f" for a {role}" if role else ""
        folder_context = f" in the context of {folder_path}" if folder_path else ""
        
        # Get system message from translation resources
        system_message = get_translation("content_generation.xlsx_generation", language)
        
        # Format the main prompt with the input parameters
        prompt_template = get_translation("content_generation.excel_prompt_template", language)
        prompt = prompt_template.format(
            file_description=file_description,
            industry=industry,
            role_context=role_context,
            folder_context=folder_context,
            language=language
        )
        
        # Add style information
        style_type = get_translation(f"style_type_mapping.xlsx.{language}", language)
        style_prompt = get_translation("content_generation.excel_style_prompt", language)
        prompt += style_prompt.format(style_type=style_type)
        
        # Request date-appropriate content
        time_prompt = get_translation("content_generation.excel_time_prompt", language)
        prompt += time_prompt
        
        content = self.llm_client.get_json_completion(
            prompt=prompt,
            system_prompt=system_message,
            max_attempts=3,
            language=language
        )
        
        return content 