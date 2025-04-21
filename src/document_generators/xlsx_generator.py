#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import logging
import random
import traceback
from datetime import date, datetime, timedelta
from typing import Dict, Tuple, Any, Optional
from pathlib import Path

# Import Excel library
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Import from local modules
from src.language_utils import get_normalized_language_key, load_prompt_templates, get_date_format

def create_excel_workbook(file_path: str, file_name: str, description: str, industry: str, language: str = "en-US", role: str = None, file_examples: Dict[str, str] = None, date_range: Tuple[date, date] = None, model: str = None) -> None:
    """Create an Excel workbook with content appropriate for the file name and description"""
    try:
        wb = Workbook()
        ws = wb.active
        
        # Load language templates for UI strings
        templates = load_prompt_templates(language)
        ui_strings = templates.get("ui_strings", {})
        sheet_title = ui_strings.get("sheet_title", "Overview")
        
        ws.title = sheet_title
        
        # Add title and description
        ws['A1'] = file_name
        ws['A2'] = description
        
        # Style the title
        title_font = Font(size=14, bold=True)
        ws['A1'].font = title_font
        
        # Add basic information
        industry_label = ui_strings.get("industry_label", "Industry:")
        role_label = ui_strings.get("role_label", "Role:")
        
        ws['A4'] = industry_label
        ws['B4'] = industry
        
        if role:
            ws['A5'] = role_label
            ws['B5'] = role
        
        # Add date range if available
        row = 6
        if date_range:
            start_date_str = get_date_format(date_range[0], "prompt", language)
            end_date_str = get_date_format(date_range[1], "prompt", language)
            period_label = ui_strings.get("period_label", "Period:")
            ws[f'A{row}'] = period_label
            ws[f'B{row}'] = f"{start_date_str} - {end_date_str}"
            row += 1
        
        # Determine appropriate data structure based on file name and industry
        # Define different data templates based on common Excel use cases
        file_name_lower = file_name.lower()
        
        # Create Data sheet with appropriate columns based on file type
        data_ws = wb.create_sheet("Data")
        
        # Apply header styling
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        
        # Setup different templates based on file name keywords
        if "budget" in file_name_lower or "financial" in file_name_lower or "forecast" in file_name_lower:
            # Financial data template
            headers = ["Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"]
            
            # Categories relevant to the industry
            categories = ["Salaries", "Equipment", "Services", "Marketing", 
                          "Operations", "Travel", "Training", "Maintenance"]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add data rows
            for row, category in enumerate(categories, 2):
                data_ws.cell(row=row, column=1, value=category)
                
                # Generate random monthly values
                yearly_total = 0
                for month in range(2, 14):  # columns 2-13 are months
                    monthly_value = round(random.uniform(1000, 10000), 2)
                    data_ws.cell(row=row, column=month, value=monthly_value)
                    yearly_total += monthly_value
                
                # Add yearly total
                data_ws.cell(row=row, column=14, value=yearly_total)
                
            # Add formulas for totals
            total_row = len(categories) + 2
            data_ws.cell(row=total_row, column=1, value="Total")
            for col in range(2, 15):
                # Sum formula for the column
                formula = f"=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})"
                data_ws.cell(row=total_row, column=col, value=formula)
        
        elif "schedule" in file_name_lower or "plan" in file_name_lower or "project" in file_name_lower:
            # Project planning template
            headers = ["Task", "Assigned To", "Start Date", "End Date", "Status", "Complete %", "Notes"]
            
            # Example tasks
            tasks = [
                "Project Initiation", "Requirements Gathering", "Design Phase", 
                "Development", "Testing", "User Acceptance", "Deployment"
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Generate dates based on date range or default to relative dates
            today = datetime.now().date()
            if date_range:
                start_date, end_date = date_range
                project_days = (end_date - start_date).days
                project_start = start_date
            else:
                project_days = 90
                project_start = today - timedelta(days=30)
            
            # Add data rows
            status_options = ["Not Started", "In Progress", "Completed", "On Hold"]
            day_offset = 0
            
            for row, task in enumerate(tasks, 2):
                # Task name
                data_ws.cell(row=row, column=1, value=task)
                
                # Assigned to
                names = ["Alex Johnson", "Sam Smith", "Jordan Lee", "Taylor Martinez"]
                data_ws.cell(row=row, column=2, value=random.choice(names))
                
                # Start date (incrementing through the project timeline)
                task_start = project_start + timedelta(days=day_offset)
                data_ws.cell(row=row, column=3, value=task_start)
                
                # Task duration between 7-21 days
                duration = random.randint(7, 21)
                day_offset += duration
                
                # End date
                task_end = task_start + timedelta(days=duration)
                data_ws.cell(row=row, column=4, value=task_end)
                
                # Status
                status = random.choice(status_options)
                data_ws.cell(row=row, column=5, value=status)
                
                # Completion percentage
                if status == "Not Started":
                    completion = 0
                elif status == "Completed":
                    completion = 100
                elif status == "On Hold":
                    completion = random.randint(10, 80)
                else:  # In Progress
                    completion = random.randint(10, 90)
                
                data_ws.cell(row=row, column=6, value=completion)
                
                # Notes
                notes = [
                    f"Pending approval", 
                    f"Needs review from stakeholders",
                    f"On schedule",
                    f"Delayed due to resource constraints",
                    ""
                ]
                data_ws.cell(row=row, column=7, value=random.choice(notes))
        
        elif "inventory" in file_name_lower or "tracker" in file_name_lower:
            # Inventory template
            headers = ["Item ID", "Item Name", "Category", "Quantity", "Unit Price", 
                       "Total Value", "Supplier", "Last Updated"]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Generate items based on industry
            categories = ["Equipment", "Supplies", "Materials", "Tools", "Consumables"]
            suppliers = ["ABC Corp", "Smith Supplies", "Tech Solutions", "Global Distribution", "Local Vendor"]
            
            for row in range(2, 12):  # 10 items
                # Item ID
                item_id = f"IT-{random.randint(1000, 9999)}"
                data_ws.cell(row=row, column=1, value=item_id)
                
                # Item Name
                item_name = f"{industry} {random.choice(['Component', 'Supply', 'Tool', 'Material', 'Kit'])}"
                data_ws.cell(row=row, column=2, value=item_name)
                
                # Category
                category = random.choice(categories)
                data_ws.cell(row=row, column=3, value=category)
                
                # Quantity
                quantity = random.randint(1, 100)
                data_ws.cell(row=row, column=4, value=quantity)
                
                # Unit Price
                unit_price = round(random.uniform(10, 1000), 2)
                data_ws.cell(row=row, column=5, value=unit_price)
                
                # Total Value (formula)
                formula = f"=D{row}*E{row}"
                data_ws.cell(row=row, column=6, value=formula)
                
                # Supplier
                supplier = random.choice(suppliers)
                data_ws.cell(row=row, column=7, value=supplier)
                
                # Last Updated
                if date_range:
                    start_date, end_date = date_range
                    update_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
                else:
                    update_date = datetime.now().date() - timedelta(days=random.randint(0, 30))
                
                data_ws.cell(row=row, column=8, value=update_date)
            
            # Add summary row
            summary_row = 12
            data_ws.cell(row=summary_row, column=1, value="Total:")
            
            # Sum formula for quantity
            quantity_formula = f"=SUM(D2:D{summary_row-1})"
            data_ws.cell(row=summary_row, column=4, value=quantity_formula)
            
            # Sum formula for total value
            value_formula = f"=SUM(F2:F{summary_row-1})"
            data_ws.cell(row=summary_row, column=6, value=value_formula)
        
        else:
            # Generic data template
            headers = ["ID", "Name", "Category", "Date", "Status", "Value", "Notes"]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Generate generic data rows
            statuses = ["Active", "Pending", "Completed", "Archived"]
            categories = ["Category A", "Category B", "Category C"]
            
            for row in range(2, 12):  # 10 rows of data
                # ID
                data_ws.cell(row=row, column=1, value=f"ID-{row-1:03d}")
                
                # Name
                data_ws.cell(row=row, column=2, value=f"Sample {industry} Item {row-1}")
                
                # Category
                data_ws.cell(row=row, column=3, value=random.choice(categories))
                
                # Date
                if date_range:
                    start_date, end_date = date_range
                    item_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
                else:
                    item_date = datetime.now().date() - timedelta(days=random.randint(0, 90))
                
                data_ws.cell(row=row, column=4, value=item_date)
                
                # Status
                data_ws.cell(row=row, column=5, value=random.choice(statuses))
                
                # Value
                data_ws.cell(row=row, column=6, value=round(random.uniform(100, 5000), 2))
                
                # Notes
                data_ws.cell(row=row, column=7, value=f"Sample notes for item {row-1}")
        
        # Auto-adjust column widths in both sheets
        for ws in [ws, data_ws]:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 30)
        
        # Save the workbook
        final_path = os.path.join(file_path, file_name)
        wb.save(final_path)
        logging.info(f"Created Excel workbook: {final_path}")
        return True
    except Exception as e:
        logging.critical(f"Error creating Excel workbook {file_path}: {e}")
        logging.critical(f"Error details: {traceback.format_exc()}")
        # Create a simple fallback file instead of exiting
        try:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = file_name
            ws['A2'] = description
            ws['A3'] = f"Industry: {industry}"
            final_path = os.path.join(file_path, file_name)
            wb.save(final_path)
            logging.warning(f"Created simple fallback Excel file due to error: {final_path}")
            return True
        except Exception as e2:
            logging.critical(f"Error creating fallback Excel file: {e2}")
            return False 